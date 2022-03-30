from openpyxl import load_workbook
import sqlite3
import sys


def convert(xlsx_filename, sqlite_filename):
    # Get Workbook
    print("Opening XLSX")
    wb = load_workbook(filename=xlsx_filename, read_only=True)
    ws = wb.active

    # Get column headings, work out details.
    print("Getting Headings")
    ws1 = ws[1]
    headings = []
    for i in range(0, len(ws1)):
        headings.append({
            'name': ws1[i].value,
            'types': [],
            'type': 'text',
        })

    # Look at some data to find column types
    print("Getting type of Headings")
    for row_idx in range(3, 50):
        row = ws[row_idx]
        for heading_idx, heading in enumerate(headings):
            if str(row[heading_idx].value).isdigit():
                heading['types'].append('integer')
            elif row[heading_idx].value:
                heading['types'].append('text')

    for heading in headings:
        types = list(set(heading['types']))
        if len(types) == 1:
            if types[0] == 'integer':
                heading['type'] = 'integer'
        heading['types'] = None

    # Create database!
    print("Create Database")
    con = sqlite3.connect(sqlite_filename)
    cur = con.cursor()

    cur.execute(
        "CREATE TABLE data (" +
        ",".join([str(h['name']) + " " + h['type'] for h in headings]) +
        ")"
    )
    con.commit()

    # data into database!
    print("Writing Data into Database")
    sql = "insert into data values (" + ",".join(["?" for h in headings]) + ")"
    for i in range(3, ws.max_row + 1):
        row = ws[i]
        data_to_insert = [c.value for c in row]
        cur.execute(sql,data_to_insert)
        con.commit()

    # Done
    con.close()


if __name__ == "__main__":
    convert(sys.argv[1], sys.argv[2])
